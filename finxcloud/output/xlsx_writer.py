"""Excel (XLSX) output writer for FinXCloud cost optimization reports.

Requires the ``openpyxl`` library (optional dependency).
Install with: ``pip install 'finxcloud[xlsx]'``
"""

from __future__ import annotations

import io
import logging
from typing import Any

log = logging.getLogger(__name__)

__all__ = ["XLSXWriter"]

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


class XLSXWriter:
    """Generate Excel reports from scan results."""

    # Style constants
    _HEADER_FONT = Font(bold=True, color="FFFFFF", size=11) if _HAS_OPENPYXL else None
    _HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") if _HAS_OPENPYXL else None
    _HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True) if _HAS_OPENPYXL else None
    _MONEY_FMT = '#,##0.00'

    @staticmethod
    def build_report_bytes(scan_result: dict) -> bytes:
        """Create a multi-sheet XLSX workbook and return it as bytes.

        Sheets:
            1. Summary — high-level metrics
            2. Recommendations — all findings
            3. Resources — resource inventory
            4. Cost Breakdown — by service and region

        Args:
            scan_result: Full scan result dict (as returned by the scan API).

        Returns:
            XLSX file content as bytes.

        Raises:
            ImportError: If openpyxl is not installed.
        """
        if not _HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install 'finxcloud[xlsx]'"
            )

        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = "Summary"
        _build_summary_sheet(ws_summary, scan_result)

        ws_recs = wb.create_sheet("Recommendations")
        _build_recommendations_sheet(ws_recs, scan_result)

        ws_resources = wb.create_sheet("Resources")
        _build_resources_sheet(ws_resources, scan_result)

        ws_cost = wb.create_sheet("Cost Breakdown")
        _build_cost_breakdown_sheet(ws_cost, scan_result)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def _style_header_row(ws, headers: list[str]) -> None:
    """Write and style a header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = XLSXWriter._HEADER_FONT
        cell.fill = XLSXWriter._HEADER_FILL
        cell.alignment = XLSXWriter._HEADER_ALIGN
    ws.freeze_panes = "A2"


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells)
        width = max(min_width, min(length + 2, max_width))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def _build_summary_sheet(ws, scan_result: dict) -> None:
    """Populate the Summary sheet with high-level metrics."""
    summary = scan_result.get("summary", {})
    overview = summary.get("overview", {})

    metrics = [
        ("Total Resources", overview.get("total_resources", 0)),
        ("30-Day Cost", overview.get("total_cost_30d", 0)),
        ("Potential Savings", overview.get("total_potential_savings", 0)),
        ("Savings %", overview.get("savings_percentage", 0)),
        ("Quick Wins", summary.get("quick_wins_count", 0)),
    ]

    _style_header_row(ws, ["Metric", "Value"])

    for row_idx, (metric, value) in enumerate(metrics, 2):
        ws.cell(row=row_idx, column=1, value=metric)
        cell = ws.cell(row=row_idx, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = XLSXWriter._MONEY_FMT

    _auto_width(ws)


def _build_recommendations_sheet(ws, scan_result: dict) -> None:
    """Populate the Recommendations sheet."""
    headers = ["Category", "Title", "Description", "Resource ID", "Region", "Effort", "Savings", "Action"]
    _style_header_row(ws, headers)

    recommendations = scan_result.get("recommendations", [])
    for row_idx, rec in enumerate(recommendations, 2):
        ws.cell(row=row_idx, column=1, value=rec.get("category", ""))
        ws.cell(row=row_idx, column=2, value=rec.get("title", ""))
        ws.cell(row=row_idx, column=3, value=rec.get("description", ""))
        ws.cell(row=row_idx, column=4, value=rec.get("resource_id", ""))
        ws.cell(row=row_idx, column=5, value=rec.get("region", ""))
        ws.cell(row=row_idx, column=6, value=rec.get("effort_level", ""))
        savings_cell = ws.cell(row=row_idx, column=7, value=rec.get("estimated_monthly_savings", 0))
        savings_cell.number_format = XLSXWriter._MONEY_FMT
        ws.cell(row=row_idx, column=8, value=rec.get("action", ""))

    _auto_width(ws)


def _build_resources_sheet(ws, scan_result: dict) -> None:
    """Populate the Resources sheet."""
    headers = ["Resource ID", "Type", "Region", "State", "Account ID"]
    _style_header_row(ws, headers)

    resources = scan_result.get("resources", [])
    for row_idx, res in enumerate(resources, 2):
        ws.cell(row=row_idx, column=1, value=res.get("resource_id", ""))
        ws.cell(row=row_idx, column=2, value=res.get("type", res.get("resource_type", "")))
        ws.cell(row=row_idx, column=3, value=res.get("region", ""))
        ws.cell(row=row_idx, column=4, value=res.get("state", ""))
        ws.cell(row=row_idx, column=5, value=res.get("account_id", ""))

    _auto_width(ws)


def _build_cost_breakdown_sheet(ws, scan_result: dict) -> None:
    """Populate the Cost Breakdown sheet (by service and region)."""
    headers = ["Service/Region", "Amount"]
    _style_header_row(ws, headers)

    cost_data = scan_result.get("cost_data", {})
    row_idx = 2

    for entry in cost_data.get("by_service", []):
        ws.cell(row=row_idx, column=1, value=entry.get("service", entry.get("name", "")))
        amount_cell = ws.cell(row=row_idx, column=2, value=entry.get("amount", 0))
        amount_cell.number_format = XLSXWriter._MONEY_FMT
        row_idx += 1

    for entry in cost_data.get("by_region", []):
        ws.cell(row=row_idx, column=1, value=entry.get("region", entry.get("name", "")))
        amount_cell = ws.cell(row=row_idx, column=2, value=entry.get("amount", 0))
        amount_cell.number_format = XLSXWriter._MONEY_FMT
        row_idx += 1

    _auto_width(ws)
